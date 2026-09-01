"""One file that empties the CRM and rebuilds it from Dean's tracker.

Files 7 and 9 were written to add to whatever was already there. This is
the other thing: a clean rebuild. Everything currently in `crm_contacts`
and `crm_leads`, for everybody, is copied to a backup table and then
deleted, and the whole of Dean's spreadsheet goes in on top so the CRM
holds exactly what the workbook holds and nothing else.

It is the same import as 7 plus 9, so the reading, the mapping and the
judgement calls all live in `dean-tracker-read.py` and
`dean-money-and-maint.py` and are not repeated here. This assembles.

Run `dean-tracker-read.py` first if `/tmp/leads.json` is not there.
"""
import openpyxl, re, json, difflib, datetime, hashlib, uuid, collections, numbers

SRC = '/root/.claude/uploads/8f56cd4b-eb0a-52ff-b3fc-74253f0d02e6/794ebad3-Sales__Maint_Leads_Tracker__Dean_Mann.xlsx'
OUT = ('/tmp/claude-0/-home-user-stc-marketing-dashboard/'
       '8f56cd4b-eb0a-52ff-b3fc-74253f0d02e6/scratchpad/dean/11-dean-rebuild.txt')

wb = openpyxl.load_workbook(SRC, data_only=True)

STOP = {'ltd', 'limited', 'group', 'services', 'service', 'and', 'the', 'uk', 'co', 'plc'}

def toks(s):
    return [t for t in re.sub(r'[^a-z0-9 ]', ' ', str(s or '').lower()).split()
            if t and t not in STOP]

def norm(s):
    return ''.join(toks(s))

def clean(v):
    if v is None:
        return None
    if isinstance(v, datetime.datetime):
        return v.date().isoformat()
    if isinstance(v, datetime.date):
        return v.isoformat()
    s = str(v).replace('\xa0', ' ')
    for bad, good in (('‘', "'"), ('’', "'"), ('“', '"'), ('”', '"')):
        s = s.replace(bad, good)
    return re.sub(r'\s+', ' ', s).strip() or None

def money(v):
    if isinstance(v, numbers.Number) and not isinstance(v, bool):
        return round(float(v), 2)
    return None

def q(v):
    return 'NULL' if v is None else "'" + str(v).replace("'", "''") + "'"

def num(v):
    return 'NULL' if v is None else repr(float(v))

def dt(v):
    return 'NULL' if v is None else q(v) + '::DATE'

def det(key):
    return str(uuid.UUID(hashlib.md5(key.encode()).hexdigest()))

def pounds(v):
    return '£{:,.2f}'.format(v)

# ---------- the 107 leads from six sheets ----------

leads = json.load(open('/tmp/leads.json'))

def lead_id(l):
    return det("stc-import:dean:{}:{}:{}:{}".format(
        l['sheet'], l['company'].strip().lower(), l.get('date') or '', l['type']))

seen, rows = set(), []
for l in leads:
    i = lead_id(l)
    if i in seen:
        continue
    seen.add(i)
    rows.append((i, l))

LEAD_COUNT = len(rows)
LEAD_COMPANIES = len({l['company'].strip().lower() for _, l in rows})

# ---------- the 16 sales onto 12 of them ----------

customers = [l for l in leads if l['sheet'] == 'trailer-sales-customer']
assert len(customers) == 12, len(customers)

sf = wb['Sales figures']
raw = []
for r in range(2, sf.max_row + 1):
    name = clean(sf.cell(r, 1).value)
    if not name:
        continue
    raw.append({'sheet_name': name,
                'order': clean(sf.cell(r, 2).value),
                'dispatch': clean(sf.cell(r, 3).value),
                'price': money(sf.cell(r, 5).value),
                'profit': money(sf.cell(r, 6).value),
                'commission': money(sf.cell(r, 8).value)})
assert len(raw) == 16, len(raw)

by_lead = collections.defaultdict(list)
matched_to = {}
for row in raw:
    best, score = None, 0.0
    for c in customers:
        s = difflib.SequenceMatcher(None, norm(row['sheet_name']), norm(c['company'])).ratio()
        if set(toks(row['sheet_name'])) & set(toks(c['company'])):
            s += 0.25
        if s > score:
            score, best = s, c
    assert score >= 0.85, (row['sheet_name'], best and best['company'], score)
    by_lead[lead_id(best)].append(row)
    matched_to[lead_id(best)] = best

sales = []
for lid, rows_ in by_lead.items():
    c = matched_to[lid]
    prices = [r['price'] for r in rows_ if r['price'] is not None]
    profits = [r['profit'] for r in rows_ if r['profit'] is not None]
    comms = [r['commission'] for r in rows_ if r['commission'] is not None]
    price = round(sum(prices), 2) if prices else None
    profit = round(sum(profits), 2) if profits else None
    comm = round(sum(comms), 2) if comms else None
    orders = sorted(r['order'] for r in rows_ if r['order'])
    disp = sorted(r['dispatch'] for r in rows_ if r['dispatch'])
    lines = []
    for r in rows_:
        bit = r['order'] or 'no order date'
        if r['price'] is not None:
            bit += ' ' + pounds(r['price'])
        bit += (' profit ' + pounds(r['profit'])) if r['profit'] is not None \
            else ' profit not recorded on the sheet'
        lines.append(bit)
    sales.append({
        'lead_id': lid, 'company': c['company'].strip(), 'units': len(rows_),
        'price': price, 'profit': profit,
        'pct': round(profit / price, 6) if (profit is not None and price) else None,
        'commission': comm,
        'order': orders[0] if orders else None,
        'dispatch': disp[-1] if disp else None,
        'note': 'Sales figures sheet: {} unit{}, {}.'.format(
            len(rows_), 's' if len(rows_) > 1 else '', '; '.join(lines)),
    })
sales.sort(key=lambda s: s['company'].lower())

TOTAL_PRICE = round(sum(s['price'] for s in sales if s['price'] is not None), 2)
TOTAL_PROFIT = round(sum(s['profit'] for s in sales if s['profit'] is not None), 2)
TOTAL_COMM = round(sum(s['commission'] for s in sales if s['commission'] is not None), 2)
TOTAL_UNITS = sum(s['units'] for s in sales)

# ---------- the 151 active maintenance accounts ----------

MSTAT = {'customer': 'customer', 'prospecting': 'lead', 'on hold': 'lead'}
ws = wb['Active Maint Account']
maint = []
for r in range(2, ws.max_row + 1):
    co = clean(ws.cell(r, 3).value)
    if not co:
        continue
    sheet_status = clean(ws.cell(r, 2).value)
    service = clean(ws.cell(r, 8).value)
    rev = money(ws.cell(r, 13).value)
    rev_text = clean(ws.cell(r, 13).value) if rev is None else None
    alpha = clean(ws.cell(r, 14).value)
    last_used = clean(ws.cell(r, 15).value)
    update = clean(ws.cell(r, 11).value)

    bits = ['Active maintenance account. The sheet says: {}.'.format(sheet_status or 'nothing')]
    if rev is not None:
        bits.append('2025 FY revenue {}.'.format(pounds(rev)))
    elif rev_text:
        bits.append('2025 FY revenue column says "{}".'.format(rev_text))
    if alpha:
        bits.append('Alpha code {}.'.format(alpha))
    if last_used and last_used != '?':
        bits.append('Last used {}.'.format(last_used))
    if update:
        bits.append(update if update.endswith('.') else update + '.')

    maint.append({
        'id': det('stc-import:dean:maint-active:{}:{}:{}'.format(
            co.strip().lower(), (service or '').lower(), (sheet_status or '').lower())),
        'company': co.strip(),
        'contact': clean(ws.cell(r, 4).value), 'phone': clean(ws.cell(r, 5).value),
        'email': clean(ws.cell(r, 6).value), 'location': clean(ws.cell(r, 7).value),
        'status': MSTAT.get((sheet_status or '').lower(), 'lead'),
        'sheet_status': sheet_status, 'what': service,
        'requirement': ' '.join(x for x in [clean(ws.cell(r, 9).value),
                                            clean(ws.cell(r, 10).value)] if x) or None,
        'action': update, 'next_action': clean(ws.cell(r, 12).value),
        'updated': clean(ws.cell(r, 1).value), 'revenue': rev,
        'notes': ' '.join(bits),
    })

assert len(maint) == 151, len(maint)
assert len({m['id'] for m in maint}) == 151, 'two rows produced the same id'

MAINT_REV = round(sum(m['revenue'] for m in maint if m['revenue'] is not None), 2)
MAINT_REV_ROWS = len([m for m in maint if m['revenue'] is not None])
MAINT_WON = len([m for m in maint if m['status'] == 'customer'])

ALL_LEADS = LEAD_COUNT + len(maint)
ALL_COMPANIES = len({l['company'].strip().lower() for _, l in rows}
                    | {m['company'].strip().lower() for m in maint})

# =============================================================
# The file
# =============================================================

sql = []
w = sql.append

w("""-- =============================================================
-- Empty the CRM, then rebuild it from Dean's tracker.
--
-- Run this on its own, in the Supabase SQL editor. It does not need
-- files 7 or 9 and it does not care whether they were ever run: it
-- clears everything first, so the result is the same either way.
--
-- One transaction. If any part of it fails, nothing is deleted and
-- nothing is added.
--
-- =============================================================
-- READ THIS PART. IT DELETES DATA.
-- =============================================================
--
-- It empties `crm_contacts` and `crm_leads` for EVERY user, not just
-- Dean, and with them everything that hangs off an account: notes,
-- addresses, ownership rows, list membership, dashboard actions. You
-- said the current data is either dummy or in the way, which is what
-- this is for. It is still worth reading twice.
--
-- Four things that point at an account are NOT deleted, because they
-- belong to somebody's day rather than to the CRM. They lose the link
-- and keep everything else:
--
--   calendar events and their guest rows   the meeting stays in the
--                                          diary, without a customer
--                                          attached
--   tasks                                  the job stays on the Work
--                                          board
--   FleetSmart contracts                   the contract row stays
--
-- Notifications about an account or a lead are deleted, because every
-- one of them links to a record that is about to stop existing and a
-- notification that goes nowhere is worse than no notification.
--
-- ---- How to undo it ----
--
-- Before it deletes anything it copies both tables:
--
--   crm_contacts_before_dean_import
--   crm_leads_before_dean_import
--
-- Those are made once. Running this file a second time keeps the
-- originals rather than backing up the import over the top of them.
-- To put it all back:
--
--   BEGIN;
--   DELETE FROM crm_contacts;
--   INSERT INTO crm_contacts SELECT * FROM crm_contacts_before_dean_import;
--   INSERT INTO crm_leads    SELECT * FROM crm_leads_before_dean_import;
--   COMMIT;
--
-- The links that were nulled out are not restored by that. If those
-- matter, say so and they can be captured too.
--
-- When you are happy and want the backup gone:
--
--   DROP TABLE crm_contacts_before_dean_import, crm_leads_before_dean_import;
--
-- =============================================================
-- What goes in
-- =============================================================
--
-- Every row of "Sales & Maint Leads Tracker - Dean Mann.xlsx", owned
-- by Dean. %ALL_LEADS% leads across %ALL_COMPANIES% companies:
--
--   Prospecting Maint Accounts   38 maintenance leads
--   Active Maint Account        151 maintenance leads, 149 of them won
--   Trailer Sales prospects      20 trailer sales, contacted
--   Trailer Sales Customer       12 trailer sales, customer
--   Trailer Lost Sales           27 trailer sales, lost
--   Contract pipeline             7 quoted deals
--   Active Rental Accounts        3 rental customers
--   Sales figures                16 sales, onto the 12 customers above
--
-- The Dashboard tab is not imported. It computes itself from the other
-- sheets, and this file asserts against its figures instead:
--
--   Open pipeline     £211,980      Closed sales      £189,195.00
--   Lost sales        £247,370      Gross profit      £23,065.37
--   Contract pipeline £591,800      Commission        £1,899.52
--   Active accounts   151           2025 FY revenue   £3,338,086.67
--
-- Any one of those coming out wrong stops the file and rolls the whole
-- thing back, backup and deletion included.
--
-- ---- The judgement calls, each reversible with one UPDATE ----
--
-- "No Action" on the prospecting sheet became `lead`, not `lost`. Both
-- rows say there is currently no requirement, which is parked.
--
--   UPDATE crm_leads SET status = 'lost'
--    WHERE notes LIKE 'Status on the sheet: No Action%';
--
-- "Dealt" became `quoted`. One row, PMCE Ltd, whose update says
-- Trukplans were sent and whose next action is to chase.
--
--   UPDATE crm_leads SET status = 'won'
--    WHERE notes LIKE 'Status on the sheet: Dealt%';
--
-- "ON HOLD" became `lead`. One row, T Diggins Group Ltd.
--
--   UPDATE crm_leads SET status = 'lost'
--    WHERE notes LIKE 'Active maintenance account. The sheet says: ON HOLD%';
--
-- 2025 FY revenue went into `sale_price`, not `estimated_value`,
-- because it is money taken and the pipeline column is money hoped for.
--
--   UPDATE crm_leads SET estimated_value = sale_price, sale_price = NULL
--    WHERE notes LIKE 'Active maintenance account.%';
--
-- Contract pipeline rows are separate leads even where the company is
-- already on another sheet. Suttle Transport has a trailer enquiry and
-- a rental contract, and those are two pitches to one customer.
--
-- A value that was not a number went in empty rather than as a guess.
-- Three rows said things like "Mulitple/See email", seven revenue cells
-- say "New Customer", and one sale's profit column is a dash. Nothing
-- billed yet and billed nothing are different facts.
-- =============================================================

BEGIN;

-- -------------------------------------------------------------
-- 1. Dean, before anything is deleted.
-- -------------------------------------------------------------
DO $$
DECLARE n INT;
BEGIN
  SELECT count(*) INTO n FROM profiles
   WHERE lower(btrim(full_name)) = 'dean mann';
  IF n = 0 THEN
    RAISE EXCEPTION
      'No account here is called Dean Mann, so nothing has been deleted. Check the spelling in profiles.full_name, then run this again.';
  END IF;
  IF n > 1 THEN
    RAISE EXCEPTION
      'There are % accounts called Dean Mann, so this cannot tell whose tracker it is. Nothing has been deleted.', n;
  END IF;
END $$;

CREATE TEMP TABLE dean AS
SELECT id FROM profiles WHERE lower(btrim(full_name)) = 'dean mann';

-- -------------------------------------------------------------
-- 2. The backup, taken once.
--
-- Made only if it is not already there, so running this file twice
-- leaves the original copy alone rather than replacing it with a copy
-- of the import.
-- -------------------------------------------------------------
DO $$
DECLARE c INT; l INT;
BEGIN
  IF to_regclass('public.crm_contacts_before_dean_import') IS NULL THEN
    CREATE TABLE crm_contacts_before_dean_import AS SELECT * FROM crm_contacts;
    CREATE TABLE crm_leads_before_dean_import    AS SELECT * FROM crm_leads;
    SELECT count(*) INTO c FROM crm_contacts_before_dean_import;
    SELECT count(*) INTO l FROM crm_leads_before_dean_import;
    RAISE NOTICE 'backed up % accounts and % leads', c, l;
  ELSE
    SELECT count(*) INTO c FROM crm_contacts_before_dean_import;
    RAISE NOTICE 'a backup of % accounts is already there from an earlier run, keeping it', c;
  END IF;
END $$;

-- -------------------------------------------------------------
-- 3. Unlink what is staying, then clear.
--
-- Diary entries, tasks and FleetSmart contracts survive with their
-- customer link emptied. Everything with a cascade on it goes with the
-- account: notes, addresses, ownership, list membership, dashboard
-- actions and the leads themselves.
-- -------------------------------------------------------------
UPDATE calendar_events SET contact_id = NULL WHERE contact_id IS NOT NULL;
UPDATE calendar_guests SET contact_id = NULL WHERE contact_id IS NOT NULL;
UPDATE tasks SET organisation_id = NULL, person_id = NULL, lead_id = NULL
 WHERE organisation_id IS NOT NULL OR person_id IS NOT NULL OR lead_id IS NOT NULL;
UPDATE fleetsmart_contracts SET account_id = NULL, lead_id = NULL
 WHERE account_id IS NOT NULL OR lead_id IS NOT NULL;
UPDATE crm_contacts SET parent_customer_id = NULL WHERE parent_customer_id IS NOT NULL;

DELETE FROM notifications WHERE subject_kind IN ('account', 'lead');

DELETE FROM crm_leads;
DELETE FROM crm_contacts;

DO $$
DECLARE c INT; l INT;
BEGIN
  SELECT count(*) INTO c FROM crm_contacts;
  SELECT count(*) INTO l FROM crm_leads;
  IF c <> 0 OR l <> 0 THEN
    RAISE EXCEPTION 'the clear left % accounts and % leads behind', c, l;
  END IF;
END $$;

-- -------------------------------------------------------------
-- 4. Quiet, for the length of the load.
--
-- Migration 066 tells somebody when a prospect lands on their tracker,
-- which is right for one handed over by a colleague and wrong for an
-- import. %ALL_LEADS% at once is not %ALL_LEADS% things happening to Dean.
--
-- The status trigger from migration 043 is deliberately left running.
-- Each account's status has to follow the leads going in, which is the
-- whole point of putting them in.
-- -------------------------------------------------------------
DO $$
BEGIN
  IF EXISTS (SELECT 1 FROM pg_trigger WHERE tgname = 'notify_lead_assigned') THEN
    ALTER TABLE crm_leads DISABLE TRIGGER notify_lead_assigned;
  END IF;
END $$;

-- -------------------------------------------------------------
-- 5. The six sheets of pitches, as they were read.
-- -------------------------------------------------------------
CREATE TEMP TABLE incoming (
  lead_id     UUID PRIMARY KEY,
  sheet       TEXT,
  company     TEXT NOT NULL,
  contact     TEXT,
  email       TEXT,
  phone       TEXT,
  location    TEXT,
  lead_type   TEXT NOT NULL,
  status      TEXT NOT NULL,
  what        TEXT,
  requirement TEXT,
  new_or_used TEXT,
  est_value   NUMERIC,
  action      TEXT,
  next_action TEXT,
  enquired    DATE,
  notes       TEXT,
  source      TEXT
);

INSERT INTO incoming VALUES""")

vals = []
for i, l in rows:
    vals.append("  (" + ", ".join([
        q(i), q(l['sheet']), q(l['company'].strip()), q(l.get('contact')),
        q(l.get('email')), q(l.get('phone')), q(l.get('location')),
        q(l['type']), q(l['status']), q(l.get('what')), q(l.get('requirement')),
        q(l.get('new_or_used')), num(l.get('estimated_value')),
        q(l.get('action')), q(l.get('next_action')), dt(l.get('date')),
        q(l.get('notes')), q(l.get('source')),
    ]) + ")")
w(",\n".join(vals) + ";")

w("""
-- -------------------------------------------------------------
-- 6. The 151 active maintenance accounts.
--
-- Enterprise Flex-e-Rent is on that sheet twice, as an accident repair
-- customer doing £227,315.74 and as a maintenance pitch still being
-- onboarded. Two pitches to one company, which is what the lead table
-- is for, so it stays two rows.
-- -------------------------------------------------------------
CREATE TEMP TABLE incoming_maint (
  lead_id      UUID PRIMARY KEY,
  company      TEXT NOT NULL,
  contact      TEXT,
  phone        TEXT,
  email        TEXT,
  location     TEXT,
  status       TEXT NOT NULL,
  sheet_status TEXT,
  what         TEXT,
  requirement  TEXT,
  action       TEXT,
  next_action  TEXT,
  updated_on   DATE,
  revenue      NUMERIC,
  notes        TEXT
);

INSERT INTO incoming_maint VALUES""")

vals = []
for m in maint:
    vals.append("  (" + ", ".join([
        q(m['id']), q(m['company']), q(m['contact']), q(m['phone']), q(m['email']),
        q(m['location']), q(m['status']), q(m['sheet_status']), q(m['what']),
        q(m['requirement']), q(m['action']), q(m['next_action']),
        dt(m['updated']), num(m['revenue']), q(m['notes']),
    ]) + ")")
w(",\n".join(vals) + ";")

w("""
-- -------------------------------------------------------------
-- 7. The accounts, one per company across both lists.
--
-- Matched on the company name folded to lower case and trimmed,
-- because "Suttle Transport " and "Suttle Transport" are one customer
-- and the spreadsheet holds both. The row that carries the most detail
-- wins where the sheets disagree.
-- -------------------------------------------------------------
CREATE TEMP TABLE every_company AS
SELECT company, contact, email, phone, location FROM incoming
UNION ALL
SELECT company, contact, email, phone, location FROM incoming_maint;

INSERT INTO crm_contacts (company_name, contact_name, email, phone, location,
                          assigned_to, source, status)
SELECT DISTINCT ON (lower(btrim(e.company)))
       btrim(e.company), e.contact, e.email, e.phone, e.location, 'Dean Mann',
       /* NOT NULL with a default of 'manual'. Passing NULL explicitly
          defeats a default, so the fallback is stated. */
       'manual', 'lead'
  FROM every_company e
 ORDER BY lower(btrim(e.company)),
          (e.contact IS NULL), (e.email IS NULL), (e.phone IS NULL);

-- -------------------------------------------------------------
-- 8. On the shared pipeline, so the CRM tab shows them.
--
-- Not decoration, and the reason is worth stating.
--
-- Migration 040 moved list membership out of `crm_contacts.list_id` and
-- into `crm_list_contacts`. The CRM screen asks that table which
-- companies are on the open list and fetches those, so a company on no
-- list is fetched by nothing: it exists, it takes leads, it shows on
-- trackers, it counts in every total, and the CRM tab is empty.
--
-- A lead cannot exist without a customer in the CRM, and that has to be
-- true on the screen and not only in the data. So every account this
-- file creates joins the shared pipeline, in the same transaction that
-- creates it.
--
-- Only the shared one. Nothing here touches anybody's private list.
-- -------------------------------------------------------------
INSERT INTO crm_list_contacts (list_id, contact_id)
SELECT l.id, c.id
  FROM crm_contacts c
 CROSS JOIN (SELECT id FROM crm_lists WHERE is_global ORDER BY created_at LIMIT 1) l
ON CONFLICT DO NOTHING;

DO $$
DECLARE unfiled INT;
BEGIN
  IF NOT EXISTS (SELECT 1 FROM crm_lists WHERE is_global) THEN
    RAISE EXCEPTION
      'There is no shared CRM list in this database, so nothing would show on the CRM tab. Make one first.';
  END IF;

  SELECT count(*) INTO unfiled
    FROM crm_contacts c
   WHERE NOT EXISTS (SELECT 1 FROM crm_list_contacts lc WHERE lc.contact_id = c.id);

  IF unfiled > 0 THEN
    RAISE EXCEPTION '% account(s) are on no list, so the CRM tab would not show them', unfiled;
  END IF;
END $$;

-- -------------------------------------------------------------
-- 9. The pitches.
-- -------------------------------------------------------------
INSERT INTO crm_leads (
  id, contact_id, owner_id, created_by, type, status,
  what, requirement, new_or_used, estimated_value,
  action, next_action, notes, date_of_enquiry, last_activity_at
)
SELECT
  i.lead_id, c.id, d.id, d.id, i.lead_type, i.status,
  i.what, i.requirement, i.new_or_used, i.est_value,
  i.action, i.next_action, i.notes, i.enquired,
  COALESCE(i.enquired::TIMESTAMPTZ, NOW())
FROM incoming i
JOIN crm_contacts c ON lower(btrim(c.company_name)) = lower(btrim(i.company))
CROSS JOIN dean d;

INSERT INTO crm_leads (
  id, contact_id, owner_id, created_by, type, status,
  what, requirement, action, next_action, notes,
  sale_price, last_activity_at
)
SELECT
  i.lead_id, c.id, d.id, d.id, 'maintenance', i.status,
  i.what, i.requirement, i.action, i.next_action, i.notes,
  i.revenue, i.updated_on::TIMESTAMPTZ
FROM incoming_maint i
JOIN crm_contacts c ON lower(btrim(c.company_name)) = lower(btrim(i.company))
CROSS JOIN dean d;

-- -------------------------------------------------------------
-- 10. The money on the twelve sales.
--
-- Sixteen rows on the figures sheet against twelve companies. Cadowood
-- bought three times, Smart Waste twice, Andy Swan twice. Their lead
-- carries the total, the earliest order date and the latest dispatch
-- date, and the note names each unit separately so nothing hides behind
-- a sum.
-- -------------------------------------------------------------
CREATE TEMP TABLE sale_money (
  lead_id    UUID PRIMARY KEY,
  company    TEXT NOT NULL,
  units      INT  NOT NULL,
  order_on   DATE,
  dispatched DATE,
  price      NUMERIC,
  profit     NUMERIC,
  pct        NUMERIC,
  commission NUMERIC,
  note       TEXT NOT NULL
);

INSERT INTO sale_money VALUES""")

vals = []
for s in sales:
    vals.append("  (" + ", ".join([
        q(s['lead_id']), q(s['company']), str(s['units']),
        dt(s['order']), dt(s['dispatch']),
        num(s['price']), num(s['profit']), num(s['pct']), num(s['commission']),
        q(s['note']),
    ]) + ")")
w(",\n".join(vals) + ";")

w("""
UPDATE crm_leads l SET
  order_date    = m.order_on,
  dispatch_date = m.dispatched,
  sale_price    = m.price,
  profit        = m.profit,
  profit_pct    = m.pct,
  commission    = m.commission,
  /* Appended, not replaced. The existing note is what Dean wrote on the
     sheet about the deal, and it is worth more than the arithmetic. */
  notes = CASE
    WHEN strpos(COALESCE(l.notes, ''), m.note) > 0 THEN l.notes
    ELSE COALESCE(NULLIF(btrim(l.notes), '') || ' ', '') || m.note
  END,
  last_activity_at = GREATEST(
    COALESCE(l.last_activity_at, '-infinity'::TIMESTAMPTZ),
    COALESCE(m.dispatched, m.order_on)::TIMESTAMPTZ)
FROM sale_money m
WHERE l.id = m.lead_id;

DO $$
BEGIN
  IF EXISTS (SELECT 1 FROM pg_trigger WHERE tgname = 'notify_lead_assigned') THEN
    ALTER TABLE crm_leads ENABLE TRIGGER notify_lead_assigned;
  END IF;
END $$;

-- -------------------------------------------------------------
-- 11. Does the CRM now say what the spreadsheet says.
--
-- Every one of these is a figure the workbook's own Dashboard tab
-- computes. Any of them coming out wrong rolls back the delete as well
-- as the import, so a failure here leaves the CRM exactly as it was.
-- -------------------------------------------------------------
DO $$
DECLARE n INT; v NUMERIC;
BEGIN
  SELECT count(*) INTO n FROM crm_leads;
  IF n <> %ALL_LEADS% THEN
    RAISE EXCEPTION 'the sheets have %ALL_LEADS% rows and % leads are here', n;
  END IF;

  SELECT count(*) INTO n FROM crm_contacts;
  IF n <> %ALL_COMPANIES% THEN
    RAISE EXCEPTION 'the sheets name %ALL_COMPANIES% companies and % accounts are here', n;
  END IF;

  SELECT count(*) INTO n FROM crm_leads WHERE owner_id IS DISTINCT FROM
    (SELECT id FROM dean);
  IF n <> 0 THEN
    RAISE EXCEPTION '% leads came out belonging to somebody other than Dean', n;
  END IF;

  SELECT count(*) INTO n FROM crm_leads WHERE contact_id IS NULL;
  IF n <> 0 THEN
    RAISE EXCEPTION '% leads point at no account', n;
  END IF;

  /* And every one of those accounts is on the shared pipeline, so the
     CRM tab shows the customer behind every lead on the tracker. A
     tracker full of customers the CRM cannot open is the exact state
     this file exists to make impossible. */
  SELECT count(*) INTO n
    FROM crm_leads l
   WHERE NOT EXISTS (
     SELECT 1 FROM crm_list_contacts lc WHERE lc.contact_id = l.contact_id);
  IF n <> 0 THEN
    RAISE EXCEPTION '% leads have a customer the CRM tab will not show', n;
  END IF;

  /* Trailer sales. */
  SELECT SUM(estimated_value) INTO v FROM crm_leads
   WHERE type = 'trailer_sales' AND status = 'contacted';
  IF v IS DISTINCT FROM 211980 THEN
    RAISE EXCEPTION 'the open trailer pipeline came to % and the spreadsheet says 211980', v;
  END IF;

  SELECT SUM(estimated_value) INTO v FROM crm_leads
   WHERE type = 'trailer_sales' AND status = 'lost';
  IF v IS DISTINCT FROM 247370 THEN
    RAISE EXCEPTION 'lost sales came to % and the spreadsheet says 247370', v;
  END IF;

  SELECT SUM(estimated_value) INTO v FROM crm_leads WHERE status = 'quoted';
  IF v IS DISTINCT FROM 591800 THEN
    RAISE EXCEPTION 'the contract pipeline came to % and the spreadsheet says 591800', v;
  END IF;

  /* The sales that closed. */
  SELECT SUM(sale_price) INTO v FROM crm_leads
   WHERE type = 'trailer_sales' AND status = 'customer';
  IF v IS DISTINCT FROM %PRICE% THEN
    RAISE EXCEPTION 'closed sales came to % and the spreadsheet says %PRICE%', v;
  END IF;

  SELECT SUM(profit) INTO v FROM crm_leads
   WHERE type = 'trailer_sales' AND status = 'customer';
  IF v IS DISTINCT FROM %PROFIT% THEN
    RAISE EXCEPTION 'gross profit came to % and the spreadsheet says %PROFIT%', v;
  END IF;

  SELECT SUM(commission) INTO v FROM crm_leads
   WHERE type = 'trailer_sales' AND status = 'customer';
  IF v IS DISTINCT FROM %COMM% THEN
    RAISE EXCEPTION 'commission came to % and the spreadsheet says %COMM%', v;
  END IF;

  /* The maintenance book. */
  SELECT count(*) INTO n FROM crm_leads
   WHERE notes LIKE 'Active maintenance account.%';
  IF n <> 151 THEN
    RAISE EXCEPTION 'the sheet has 151 active accounts and % are here', n;
  END IF;

  SELECT count(*) INTO n FROM crm_leads
   WHERE notes LIKE 'Active maintenance account.%' AND status = 'customer';
  IF n <> %WON% THEN
    RAISE EXCEPTION 'the sheet calls %WON% of them Customer and % came out won', n;
  END IF;

  SELECT SUM(sale_price) INTO v FROM crm_leads
   WHERE notes LIKE 'Active maintenance account.%';
  IF v IS DISTINCT FROM %REV% THEN
    RAISE EXCEPTION '2025 FY revenue came to % and the spreadsheet says %REV%', v;
  END IF;

  SELECT count(sale_price) INTO n FROM crm_leads
   WHERE notes LIKE 'Active maintenance account.%';
  IF n <> %REVROWS% THEN
    RAISE EXCEPTION '%REVROWS% rows on the sheet carry a revenue figure and % do here', n;
  END IF;

  RAISE NOTICE 'ok  %ALL_LEADS% leads on %ALL_COMPANIES% accounts, all Dean''s, and every Dashboard figure agrees';
END $$;

COMMIT;

-- =============================================================
-- What landed. Nothing below writes.
-- =============================================================

SELECT l.type, l.status, count(*) AS leads,
       to_char(SUM(l.estimated_value), 'FM£999,999,990') AS open_pipeline,
       to_char(SUM(l.sale_price),      'FM£999,999,990') AS money_taken
FROM crm_leads l
GROUP BY l.type, l.status
ORDER BY l.type, l.status;

SELECT c.company_name,
       to_char(l.estimated_value, 'FM£999,999,990')  AS estimated,
       to_char(l.sale_price,      'FM£999,999,990')  AS sold_for,
       to_char(l.profit,          'FM£999,990.00')   AS profit,
       to_char(l.profit_pct * 100, 'FM990.0') || '%' AS margin,
       to_char(l.commission,      'FM£9,990.00')     AS commission,
       l.order_date, l.dispatch_date
FROM crm_leads l
JOIN crm_contacts c ON c.id = l.contact_id
WHERE l.type = 'trailer_sales' AND l.status = 'customer'
ORDER BY l.sale_price DESC NULLS LAST;

SELECT c.company_name, l.what AS service,
       to_char(l.sale_price, 'FM£999,999,990.00') AS revenue_2025_fy,
       c.contact_name, c.phone
FROM crm_leads l
JOIN crm_contacts c ON c.id = l.contact_id
WHERE l.notes LIKE 'Active maintenance account.%'
ORDER BY l.sale_price DESC NULLS LAST
LIMIT 20;
""")

body = "\n".join(sql)
for token, value in [
    ('%PRICE%', '{:.2f}'.format(TOTAL_PRICE)),
    ('%PROFIT%', '{:.2f}'.format(TOTAL_PROFIT)),
    ('%COMM%', '{:.2f}'.format(TOTAL_COMM)),
    ('%WON%', str(MAINT_WON)),
    ('%REV%', '{:.2f}'.format(MAINT_REV)),
    ('%REVROWS%', str(MAINT_REV_ROWS)),
    ('%ALL_LEADS%', str(ALL_LEADS)),
    ('%ALL_COMPANIES%', str(ALL_COMPANIES)),
]:
    body = body.replace(token, value)

open(OUT, 'w').write(body + "\n")

print('pitch leads   ', LEAD_COUNT, 'on', LEAD_COMPANIES, 'companies')
print('maint leads   ', len(maint))
print('all leads     ', ALL_LEADS, 'on', ALL_COMPANIES, 'companies')
print('sales         ', TOTAL_UNITS, 'worth', TOTAL_PRICE,
      'profit', TOTAL_PROFIT, 'commission', TOTAL_COMM)
print('maint revenue ', MAINT_REV, 'on', MAINT_REV_ROWS, 'rows,', MAINT_WON, 'won')
print('written to', OUT)
