"""The second half of Dean's tracker: the money on the sales, and the
active maintenance accounts.

`dean-tracker-read.py` and `dean-tracker-sql.py` produced file 7, which
put 107 leads in from six sheets. Two things were left out of that file
and both of them are needed for the CRM to say what the spreadsheet
says:

  Sales figures         16 rows of real money against 12 companies that
                        went in as customer leads with an estimate on
                        them and nothing else. An estimate is what
                        somebody hoped for. These are what happened.

  Active Maint Account  151 rows. The accounts were already in the CRM
                        as companies, so nothing was pitching to them
                        and nothing said they were won. On the tracker
                        they did not exist at all.

This writes file 9. It reads the workbook directly for the money and
reuses `/tmp/leads.json` only to recover the ids file 7 gave the twelve
customer leads, so the update lands on the same rows rather than on
whatever the company name matches today.

Run `dean-tracker-read.py` first if `/tmp/leads.json` is not there. It
is what writes it.
"""
import openpyxl, re, json, difflib, datetime, hashlib, uuid, collections, numbers

SRC = '/root/.claude/uploads/8f56cd4b-eb0a-52ff-b3fc-74253f0d02e6/794ebad3-Sales__Maint_Leads_Tracker__Dean_Mann.xlsx'
OUT = ('/tmp/claude-0/-home-user-stc-marketing-dashboard/'
       '8f56cd4b-eb0a-52ff-b3fc-74253f0d02e6/scratchpad/dean/9-dean-money-and-maint.txt')

wb = openpyxl.load_workbook(SRC, data_only=True)

# ---------- shared bits ----------

STOP = {'ltd', 'limited', 'group', 'services', 'service', 'and', 'the', 'uk', 'co', 'plc'}

def toks(s):
    return [t for t in re.sub(r'[^a-z0-9 ]', ' ', str(s or '').lower()).split()
            if t and t not in STOP]

def norm(s):
    return ''.join(toks(s))

def clean(v):
    if v is None:
        return None
    if isinstance(v, datetime.datetime):
        return v.date().isoformat()
    if isinstance(v, datetime.date):
        return v.isoformat()
    s = str(v).replace('\xa0', ' ')
    # Word's curly quotes, flattened. They survive a paste into the SQL
    # editor and they survive Postgres, but they are one more thing that
    # can go wrong in a file somebody is going to paste, and nothing is
    # lost by writing an apostrophe as an apostrophe.
    for bad, good in (('‘', "'"), ('’', "'"), ('“', '"'), ('”', '"')):
        s = s.replace(bad, good)
    s = re.sub(r'\s+', ' ', s).strip()
    return s or None

def money(v):
    """A number, or nothing.

    Seven revenue cells say "New Customer" and one profit column says
    "-". Those are words, not amounts, and turning them into zero would
    put a company that has not been billed yet next to a company that
    was billed nothing, which are different facts.
    """
    if isinstance(v, numbers.Number) and not isinstance(v, bool):
        return round(float(v), 2)
    return None

def q(v):
    return 'NULL' if v is None else "'" + str(v).replace("'", "''") + "'"

def num(v):
    return 'NULL' if v is None else repr(float(v))

def dt(v):
    return 'NULL' if v is None else q(v) + '::DATE'

def det(key):
    return str(uuid.UUID(hashlib.md5(key.encode()).hexdigest()))

def pounds(v):
    return '£{:,.2f}'.format(v)

# =============================================================
# A. The sixteen sales, aggregated onto twelve leads.
# =============================================================

leads = json.load(open('/tmp/leads.json'))

def lead_id(l):
    key = "stc-import:dean:{}:{}:{}:{}".format(
        l['sheet'], l['company'].strip().lower(), l.get('date') or '', l['type'])
    return det(key)

customers = [l for l in leads if l['sheet'] == 'trailer-sales-customer']
assert len(customers) == 12, len(customers)

sf = wb['Sales figures']
raw = []
for r in range(2, sf.max_row + 1):
    name = clean(sf.cell(r, 1).value)
    if not name:
        continue
    raw.append({
        'sheet_name': name,
        'order': clean(sf.cell(r, 2).value),
        'dispatch': clean(sf.cell(r, 3).value),
        'price': money(sf.cell(r, 5).value),
        'profit': money(sf.cell(r, 6).value),
        'commission': money(sf.cell(r, 8).value),
    })
assert len(raw) == 16, len(raw)

# Matched on the company name rather than on a key, because the figures
# sheet has no key column. The threshold is deliberately high and the
# assert is deliberately fatal: a sale attached to the wrong customer is
# worse than a sale left off, and a near miss should stop the file being
# written rather than be quietly resolved to the closest thing.
by_lead = collections.defaultdict(list)
matched_to = {}
for row in raw:
    best, score = None, 0.0
    for c in customers:
        s = difflib.SequenceMatcher(None, norm(row['sheet_name']), norm(c['company'])).ratio()
        if set(toks(row['sheet_name'])) & set(toks(c['company'])):
            s += 0.25
        if s > score:
            score, best = s, c
    assert score >= 0.85, (row['sheet_name'], best and best['company'], score)
    by_lead[lead_id(best)].append(row)
    matched_to[lead_id(best)] = best

sales = []
for lid, rows_ in by_lead.items():
    c = matched_to[lid]
    prices = [r['price'] for r in rows_ if r['price'] is not None]
    profits = [r['profit'] for r in rows_ if r['profit'] is not None]
    comms = [r['commission'] for r in rows_ if r['commission'] is not None]
    price = round(sum(prices), 2) if prices else None
    profit = round(sum(profits), 2) if profits else None
    comm = round(sum(comms), 2) if comms else None
    orders = sorted(r['order'] for r in rows_ if r['order'])
    disp = sorted(r['dispatch'] for r in rows_ if r['dispatch'])

    lines = []
    for r in rows_:
        bit = r['order'] or 'no order date'
        if r['price'] is not None:
            bit += ' ' + pounds(r['price'])
        bit += (' profit ' + pounds(r['profit'])) if r['profit'] is not None \
            else ' profit not recorded on the sheet'
        lines.append(bit)

    note = 'Sales figures sheet: {} unit{}, {}.'.format(
        len(rows_), 's' if len(rows_) > 1 else '', '; '.join(lines))

    sales.append({
        'lead_id': lid,
        'company': c['company'].strip(),
        'units': len(rows_),
        'price': price,
        'profit': profit,
        'pct': round(profit / price, 6) if (profit is not None and price) else None,
        'commission': comm,
        'order': orders[0] if orders else None,
        'dispatch': disp[-1] if disp else None,
        'note': note,
    })

sales.sort(key=lambda s: s['company'].lower())

TOTAL_PRICE = round(sum(s['price'] for s in sales if s['price'] is not None), 2)
TOTAL_PROFIT = round(sum(s['profit'] for s in sales if s['profit'] is not None), 2)
TOTAL_COMM = round(sum(s['commission'] for s in sales if s['commission'] is not None), 2)
TOTAL_UNITS = sum(s['units'] for s in sales)

# =============================================================
# B. The 151 active maintenance accounts.
# =============================================================

# The sheet's word, and what it becomes. "ON HOLD" is not one of the six
# the column allows. It becomes `lead`, which keeps the company in the
# pipeline without claiming they are paying, and the sheet's own word is
# kept in the notes so the choice is visible and reversible.
MSTAT = {'customer': 'customer', 'prospecting': 'lead', 'on hold': 'lead'}

ws = wb['Active Maint Account']
maint = []
for r in range(2, ws.max_row + 1):
    co = clean(ws.cell(r, 3).value)
    if not co:
        continue

    sheet_status = clean(ws.cell(r, 2).value)
    service = clean(ws.cell(r, 8).value)
    rev = money(ws.cell(r, 13).value)
    rev_text = clean(ws.cell(r, 13).value) if rev is None else None
    alpha = clean(ws.cell(r, 14).value)
    last_used = clean(ws.cell(r, 15).value)
    update = clean(ws.cell(r, 11).value)

    bits = ['Active maintenance account. The sheet says: {}.'.format(sheet_status or 'nothing')]
    if rev is not None:
        bits.append('2025 FY revenue {}.'.format(pounds(rev)))
    elif rev_text:
        bits.append('2025 FY revenue column says "{}".'.format(rev_text))
    if alpha:
        bits.append('Alpha code {}.'.format(alpha))
    if last_used and last_used != '?':
        bits.append('Last used {}.'.format(last_used))
    if update:
        bits.append(update if update.endswith('.') else update + '.')

    maint.append({
        'id': det('stc-import:dean:maint-active:{}:{}:{}'.format(
            co.strip().lower(), (service or '').lower(), (sheet_status or '').lower())),
        'company': co.strip(),
        'contact': clean(ws.cell(r, 4).value),
        'phone': clean(ws.cell(r, 5).value),
        'email': clean(ws.cell(r, 6).value),
        'location': clean(ws.cell(r, 7).value),
        'status': MSTAT.get((sheet_status or '').lower(), 'lead'),
        'sheet_status': sheet_status,
        'what': service,
        'requirement': ' '.join(x for x in [clean(ws.cell(r, 9).value),
                                            clean(ws.cell(r, 10).value)] if x) or None,
        'action': update,
        'next_action': clean(ws.cell(r, 12).value),
        'updated': clean(ws.cell(r, 1).value),
        'revenue': rev,
        'notes': ' '.join(bits),
    })

assert len(maint) == 151, len(maint)
assert len({m['id'] for m in maint}) == 151, 'two rows produced the same id'

MAINT_REV = round(sum(m['revenue'] for m in maint if m['revenue'] is not None), 2)
MAINT_REV_ROWS = len([m for m in maint if m['revenue'] is not None])
MAINT_WON = len([m for m in maint if m['status'] == 'customer'])
MAINT_COMPANIES = len({m['company'].strip().lower() for m in maint})

# =============================================================
# C. The file.
# =============================================================

sql = []
w = sql.append

w("""-- =============================================================
-- Dean's tracker, part two: the money and the live accounts.
--
-- Run this AFTER 7-dean-leads.txt, in the Supabase SQL editor. It
-- refuses to run if that one has not been, because everything in
-- section 2 is an update to rows that file created.
--
-- One transaction. If any part fails, none of it lands.
--
-- ---- What this is for ----
--
-- File 7 put the pitches in. It did not put in what the pitches
-- actually earned, and it left one whole sheet out. Two gaps, and both
-- of them are the difference between the CRM agreeing with the
-- spreadsheet and merely resembling it.
--
--   Sales figures         16 sales against 12 companies. Those twelve
--                         went in as customer leads carrying an
--                         estimated value and nothing else. An estimate
--                         is what somebody hoped for in March. This
--                         puts the order date, the dispatch date, the
--                         price, the profit, the margin and the
--                         commission on them.
--
--   Active Maint Account  151 rows. You said the companies were already
--                         in the CRM, and they are, but only as
--                         companies. Nothing said anybody had won them
--                         and none of them appeared on the tracker at
--                         all. This puts a maintenance lead against
--                         each one, at the status the sheet gives it.
--
-- ---- Numbers it will not let itself get wrong ----
--
-- The Dashboard tab of the workbook computes its own totals from these
-- sheets. This file asserts against them, so it stops rather than
-- landing a figure that disagrees with the spreadsheet:
--
--   Closed sales      £189,195.00
--   Gross profit      £23,065.37
--   Commission        £1,899.52
--   Active accounts   151
--   2025 FY revenue   £3,338,086.67
--
-- ---- Safe to run more than once ----
--
-- Every lead here has an id derived from its row, so the second run
-- inserts nothing. The money columns are written rather than filled in,
-- because the spreadsheet is the source of truth for them and writing
-- the same figure twice changes nothing. The note added to each sale is
-- checked for first, so it is not appended twice.
--
-- ---- Three judgement calls ----
--
-- 1. Revenue goes into `sale_price`, not `estimated_value`.
--
--    2025 FY revenue is money taken, not money hoped for, and these
--    leads are already at `customer`. Putting it in `estimated_value`
--    would add £3.3m to the open pipeline figure, which counts money
--    already earned as money still to win.
--
--    To move it, if you would rather see it in the pipeline column:
--
--      UPDATE crm_leads SET estimated_value = sale_price, sale_price = NULL
--       WHERE notes LIKE 'Active maintenance account.%';
--
-- 2. "ON HOLD" became `lead`. One row, T Diggins Group Ltd, whose
--    update says the account is set up and whose revenue is blank. Held
--    is not lost and it is not paying, so it stays in the pipeline. The
--    sheet's own word is in the notes.
--
--      UPDATE crm_leads SET status = 'lost'
--       WHERE notes LIKE 'Active maintenance account. The sheet says: ON HOLD%';
--
-- 3. Enterprise Flex-e-Rent gets two leads, not one. The sheet has it
--    twice: an accident repair account doing £227,315.74, and a
--    separate maintenance pitch still being onboarded. Those are two
--    pitches to one company, which is what the lead table is for.
--
-- ---- What is deliberately not written ----
--
-- `commission_rate`. The figures sheet pays 10% on some sales and 7.5%
-- on others, including on two sales to the same customer, so there is
-- no one rate per account to record. The commission itself is exact and
-- comes straight off the sheet.
--
-- Seven revenue cells say "New Customer" and one sale's profit column
-- says "-". Those went in empty with the sheet's own wording kept in
-- the notes, rather than as zero. Nothing billed yet and billed nothing
-- are different facts.
-- =============================================================

BEGIN;

-- -------------------------------------------------------------
-- 1. Dean, and file 7.
-- -------------------------------------------------------------
DO $$
DECLARE n INT;
BEGIN
  SELECT count(*) INTO n FROM profiles
   WHERE lower(btrim(full_name)) = 'dean mann';
  IF n = 0 THEN
    RAISE EXCEPTION
      'No account here is called Dean Mann. Check the spelling in profiles.full_name, then run this again.';
  END IF;
  IF n > 1 THEN
    RAISE EXCEPTION
      'There are % accounts called Dean Mann, so this cannot tell whose tracker it is.', n;
  END IF;
END $$;

CREATE TEMP TABLE dean AS
SELECT id FROM profiles WHERE lower(btrim(full_name)) = 'dean mann';

-- -------------------------------------------------------------
-- 2. The sales.
--
-- Sixteen rows on the figures sheet, aggregated onto the twelve leads
-- file 7 created. Three companies bought more than once: Cadowood
-- three times, Smart Waste twice, Andy Swan twice. Their lead carries
-- the total, the earliest order date, the latest dispatch date, and a
-- note naming each unit separately so nothing is hidden behind a sum.
-- -------------------------------------------------------------
CREATE TEMP TABLE sale_money (
  lead_id    UUID PRIMARY KEY,
  company    TEXT NOT NULL,
  units      INT  NOT NULL,
  order_on   DATE,
  dispatched DATE,
  price      NUMERIC,
  profit     NUMERIC,
  pct        NUMERIC,
  commission NUMERIC,
  note       TEXT NOT NULL
);

INSERT INTO sale_money VALUES""")

vals = []
for s in sales:
    vals.append("  (" + ", ".join([
        q(s['lead_id']), q(s['company']), str(s['units']),
        dt(s['order']), dt(s['dispatch']),
        num(s['price']), num(s['profit']), num(s['pct']), num(s['commission']),
        q(s['note']),
    ]) + ")")
w(",\n".join(vals) + ";")

w("""
DO $$
DECLARE found INT;
BEGIN
  SELECT count(*) INTO found
    FROM sale_money m JOIN crm_leads l ON l.id = m.lead_id;
  IF found <> 12 THEN
    RAISE EXCEPTION
      'Expected the 12 customer leads from file 7 and found %. Run 7-dean-leads.txt first.', found;
  END IF;
END $$;

UPDATE crm_leads l SET
  order_date    = m.order_on,
  dispatch_date = m.dispatched,
  sale_price    = m.price,
  profit        = m.profit,
  profit_pct    = m.pct,
  commission    = m.commission,
  /* Appended, not replaced. The existing note is what Dean wrote on
     the sheet about the deal, and it is worth more than the arithmetic.
     Checked for first so a second run does not append it again. */
  notes = CASE
    WHEN strpos(COALESCE(l.notes, ''), m.note) > 0 THEN l.notes
    ELSE COALESCE(NULLIF(btrim(l.notes), '') || ' ', '') || m.note
  END,
  last_activity_at = GREATEST(
    COALESCE(l.last_activity_at, '-infinity'::TIMESTAMPTZ),
    COALESCE(m.dispatched, m.order_on)::TIMESTAMPTZ)
FROM sale_money m
WHERE l.id = m.lead_id;

-- The three figures the Dashboard tab computes, asserted rather than
-- reported. A total that has drifted stops the file here.
DO $$
DECLARE p NUMERIC; g NUMERIC; c NUMERIC;
BEGIN
  SELECT SUM(l.sale_price), SUM(l.profit), SUM(l.commission)
    INTO p, g, c
    FROM crm_leads l JOIN sale_money m ON m.lead_id = l.id;

  IF p IS DISTINCT FROM %PRICE% THEN
    RAISE EXCEPTION 'Closed sales came to % and the spreadsheet says %PRICE%', p;
  END IF;
  IF g IS DISTINCT FROM %PROFIT% THEN
    RAISE EXCEPTION 'Gross profit came to % and the spreadsheet says %PROFIT%', g;
  END IF;
  IF c IS DISTINCT FROM %COMM% THEN
    RAISE EXCEPTION 'Commission came to % and the spreadsheet says %COMM%', c;
  END IF;

  RAISE NOTICE 'ok  %UNITS% sales worth % on 12 customers, profit %, commission %', p, g, c;
END $$;

-- -------------------------------------------------------------
-- 3. The active maintenance accounts.
--
-- 151 rows, %COMPANIES% companies. Enterprise Flex-e-Rent is on the
-- sheet twice, as an accident repair customer and as a maintenance
-- prospect, and stays two leads on one account.
-- -------------------------------------------------------------
CREATE TEMP TABLE incoming_maint (
  lead_id      UUID PRIMARY KEY,
  company      TEXT NOT NULL,
  contact      TEXT,
  phone        TEXT,
  email        TEXT,
  location     TEXT,
  status       TEXT NOT NULL,
  sheet_status TEXT,
  what         TEXT,
  requirement  TEXT,
  action       TEXT,
  next_action  TEXT,
  updated_on   DATE,
  revenue      NUMERIC,
  notes        TEXT
);

INSERT INTO incoming_maint VALUES""")

vals = []
for m in maint:
    vals.append("  (" + ", ".join([
        q(m['id']), q(m['company']), q(m['contact']), q(m['phone']), q(m['email']),
        q(m['location']), q(m['status']), q(m['sheet_status']), q(m['what']),
        q(m['requirement']), q(m['action']), q(m['next_action']),
        dt(m['updated']), num(m['revenue']), q(m['notes']),
    ]) + ")")
w(",\n".join(vals) + ";")

w("""
-- Any of these not already a company in the CRM. You said the active
-- accounts were imported, so this should mostly find nothing, and it is
-- here because "mostly" is not a thing to run an import on.
INSERT INTO crm_contacts (company_name, contact_name, email, phone, location,
                          assigned_to, source, status)
SELECT DISTINCT ON (lower(btrim(i.company)))
       btrim(i.company), i.contact, i.email, i.phone, i.location, 'Dean Mann',
       'manual', 'lead'
  FROM incoming_maint i
 WHERE NOT EXISTS (
   SELECT 1 FROM crm_contacts c
    WHERE lower(btrim(c.company_name)) = lower(btrim(i.company)))
 ORDER BY lower(btrim(i.company)),
          (i.contact IS NULL), (i.email IS NULL), (i.phone IS NULL);

-- Filled in where the CRM's own is empty, never overwritten. Anything
-- corrected in the CRM since the sheet was written stays corrected.
UPDATE crm_contacts c SET
  contact_name = COALESCE(NULLIF(btrim(c.contact_name), ''), f.contact),
  email        = COALESCE(NULLIF(btrim(c.email), ''),        f.email),
  phone        = COALESCE(NULLIF(btrim(c.phone), ''),        f.phone),
  location     = COALESCE(NULLIF(btrim(c.location), ''),     f.location),
  assigned_to  = COALESCE(NULLIF(btrim(c.assigned_to), ''),  'Dean Mann')
FROM (
  SELECT DISTINCT ON (lower(btrim(company)))
         lower(btrim(company)) AS key, contact, email, phone, location
    FROM incoming_maint
   ORDER BY lower(btrim(company)),
            (contact IS NULL), (email IS NULL), (phone IS NULL)
) f
WHERE lower(btrim(c.company_name)) = f.key;

-- Quiet, for the length of the load, for the same reason file 7 was.
-- 151 accounts arriving at once is not 151 things happening to Dean.
DO $$
BEGIN
  IF EXISTS (SELECT 1 FROM pg_trigger WHERE tgname = 'notify_lead_assigned') THEN
    ALTER TABLE crm_leads DISABLE TRIGGER notify_lead_assigned;
  END IF;
END $$;

INSERT INTO crm_leads (
  id, contact_id, owner_id, created_by, type, status,
  what, requirement, action, next_action, notes,
  sale_price, last_activity_at
)
SELECT
  i.lead_id, c.id, d.id, d.id, 'maintenance', i.status,
  i.what, i.requirement, i.action, i.next_action, i.notes,
  i.revenue, i.updated_on::TIMESTAMPTZ
FROM incoming_maint i
JOIN crm_contacts c ON lower(btrim(c.company_name)) = lower(btrim(i.company))
CROSS JOIN dean d
ON CONFLICT (id) DO NOTHING;

DO $$
BEGIN
  IF EXISTS (SELECT 1 FROM pg_trigger WHERE tgname = 'notify_lead_assigned') THEN
    ALTER TABLE crm_leads ENABLE TRIGGER notify_lead_assigned;
  END IF;
END $$;

-- -------------------------------------------------------------
-- 4. Did the sheet and the CRM end up saying the same thing.
-- -------------------------------------------------------------
DO $$
DECLARE made INT; won INT; rev NUMERIC; rev_rows INT;
BEGIN
  SELECT count(*) INTO made
    FROM crm_leads l JOIN incoming_maint i ON i.lead_id = l.id;
  IF made <> 151 THEN
    RAISE EXCEPTION 'The sheet has 151 active accounts and % leads are here', made;
  END IF;

  SELECT count(*) INTO won
    FROM crm_leads l JOIN incoming_maint i ON i.lead_id = l.id
   WHERE l.status = 'customer';
  IF won <> %WON% THEN
    RAISE EXCEPTION 'The sheet calls %WON% of them Customer and % came out won', won;
  END IF;

  SELECT SUM(l.sale_price), count(l.sale_price) INTO rev, rev_rows
    FROM crm_leads l JOIN incoming_maint i ON i.lead_id = l.id;
  IF rev IS DISTINCT FROM %REV% THEN
    RAISE EXCEPTION '2025 FY revenue came to % and the spreadsheet says %REV%', rev;
  END IF;
  IF rev_rows <> %REVROWS% THEN
    RAISE EXCEPTION '%REVROWS% rows on the sheet carry a revenue figure and % do here', rev_rows;
  END IF;

  RAISE NOTICE 'ok  151 active accounts, % of them won, % of revenue', won, rev;
END $$;

-- Every account that has a lead of Dean's is now assigned to him and
-- carries a status the trigger from migration 043 worked out. Nothing
-- here set that column: it is a reading of the leads that just went in.
COMMIT;

-- =============================================================
-- What landed. Nothing below writes.
-- =============================================================

-- The twelve sales, next to the estimate they were pitched at.
SELECT c.company_name,
       to_char(l.estimated_value, 'FM£999,999,990') AS estimated,
       to_char(l.sale_price,      'FM£999,999,990') AS sold_for,
       to_char(l.profit,          'FM£999,990.00')  AS profit,
       to_char(l.profit_pct * 100, 'FM990.0') || '%' AS margin,
       to_char(l.commission,      'FM£9,990.00')    AS commission,
       l.order_date, l.dispatch_date
FROM crm_leads l
JOIN crm_contacts c ON c.id = l.contact_id
WHERE l.type = 'trailer_sales' AND l.status = 'customer'
  AND l.sale_price IS NOT NULL
ORDER BY l.sale_price DESC;

-- And the totals, which should be the Dashboard tab's three figures.
SELECT count(*) AS customers,
       to_char(SUM(sale_price), 'FM£999,999,990.00') AS closed_sales,
       to_char(SUM(profit),     'FM£999,999,990.00') AS gross_profit,
       to_char(SUM(commission), 'FM£999,999,990.00') AS commission
FROM crm_leads
WHERE type = 'trailer_sales' AND status = 'customer' AND sale_price IS NOT NULL;

-- The maintenance book, by where each account stands.
SELECT l.status,
       count(*) AS accounts,
       count(l.sale_price) AS with_revenue,
       to_char(SUM(l.sale_price), 'FM£999,999,990.00') AS revenue_2025_fy
FROM crm_leads l
JOIN profiles p ON p.id = l.owner_id
WHERE lower(btrim(p.full_name)) = 'dean mann'
  AND l.type = 'maintenance'
  AND l.notes LIKE 'Active maintenance account.%'
GROUP BY l.status
ORDER BY l.status;

-- The twenty biggest, so the top of the book is visible rather than
-- summed.
SELECT c.company_name, l.what AS service, l.status,
       to_char(l.sale_price, 'FM£999,999,990.00') AS revenue_2025_fy,
       c.contact_name, c.phone
FROM crm_leads l
JOIN crm_contacts c ON c.id = l.contact_id
WHERE l.notes LIKE 'Active maintenance account.%'
ORDER BY l.sale_price DESC NULLS LAST
LIMIT 20;

-- The rows that carry no revenue figure, named rather than counted, so
-- the nine of them are a list to check and not a gap to wonder about.
SELECT c.company_name, l.status, left(l.notes, 110) AS what_the_sheet_said
FROM crm_leads l
JOIN crm_contacts c ON c.id = l.contact_id
WHERE l.notes LIKE 'Active maintenance account.%' AND l.sale_price IS NULL
ORDER BY c.company_name;
""")

body = "\n".join(sql)
for token, value in [
    ('%PRICE%', '{:.2f}'.format(TOTAL_PRICE)),
    ('%PROFIT%', '{:.2f}'.format(TOTAL_PROFIT)),
    ('%COMM%', '{:.2f}'.format(TOTAL_COMM)),
    ('%UNITS%', str(TOTAL_UNITS)),
    ('%WON%', str(MAINT_WON)),
    ('%REV%', '{:.2f}'.format(MAINT_REV)),
    ('%REVROWS%', str(MAINT_REV_ROWS)),
    ('%COMPANIES%', str(MAINT_COMPANIES)),
]:
    body = body.replace(token, value)

open(OUT, 'w').write(body + "\n")

print('sales rows       ', len(raw), 'onto', len(sales), 'leads')
print('closed sales     ', TOTAL_PRICE)
print('gross profit     ', TOTAL_PROFIT)
print('commission       ', TOTAL_COMM)
print('maint leads      ', len(maint), 'on', MAINT_COMPANIES, 'companies')
print('maint won        ', MAINT_WON)
print('maint revenue    ', MAINT_REV, 'on', MAINT_REV_ROWS, 'rows')
print('written to', OUT)
