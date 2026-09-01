"""Reading Dean Mann's tracker out of the spreadsheet.

Kept because the SQL that came out of it is 107 rows of hand unreadable
VALUES, and the only way to check the mapping is to read the thing that
made it. Pair it with `dean-tracker-sql.py`, which turns what this
produces into the file somebody pastes into Supabase.

  python3 scripts/import/dean-tracker-read.py   # writes /tmp/leads.json
  python3 scripts/import/dean-tracker-sql.py    # writes the .txt

Point SRC at wherever the workbook is. It is not in this repository:
it is somebody's working spreadsheet and it does not belong here.
"""
import openpyxl, re, datetime, json, hashlib

SRC = '794ebad3-Sales__Maint_Leads_Tracker__Dean_Mann.xlsx'
wb = openpyxl.load_workbook(SRC, data_only=True)

def clean(v):
    if v is None: return None
    if isinstance(v, datetime.datetime): return v.date().isoformat()
    if isinstance(v, datetime.date): return v.isoformat()
    s = str(v).replace(' ', ' ').strip()
    s = re.sub(r'\s+', ' ', s)
    return s or None

def money(v):
    """A number, or nothing. 'Mulitple/See email' is not a number."""
    if v is None: return None
    if isinstance(v, (int, float)): return float(v)
    s = re.sub(r'[^0-9.]', '', str(v))
    if not s or s.count('.') > 1: return None
    try:
        n = float(s)
        return n if 0 < n < 100_000_000 else None
    except ValueError:
        return None

def when(v):
    """A date, or nothing. Some cells hold '31/03/26 - Spoke to...'."""
    if isinstance(v, (datetime.datetime, datetime.date)):
        return (v.date() if isinstance(v, datetime.datetime) else v).isoformat()
    s = clean(v)
    if not s: return None
    m = re.match(r'(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})', s)
    if not m: return None
    d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
    if y < 100: y += 2000
    try: return datetime.date(y, mo, d).isoformat()
    except ValueError: return None

def rows(ws, hdr=1):
    heads = [clean(ws.cell(hdr, c).value) for c in range(1, ws.max_column + 1)]
    out = []
    for r in range(hdr + 1, ws.max_row + 1):
        d = {}
        for c in range(1, ws.max_column + 1):
            key = heads[c - 1] or f'col{c}'
            d[key] = ws.cell(r, c).value
        if any(v is not None and str(v).strip() != '' for v in d.values()):
            out.append(d)
    return out

MAINT_STATUS = {
    'prospecting': 'lead',
    'on going':    'contacted',
    'dealt':       'quoted',
    # "Currently no requirement for additional support" is parked, not
    # lost. Marking it lost would take it out of the pipeline and say
    # something the sheet does not say.
    'no action':   'lead',
}

leads = []

def add(**kw):
    kw.setdefault('new_or_used', None)
    kw.setdefault('estimated_value', None)
    leads.append(kw)

def joined(*parts):
    bits = [p for p in parts if p]
    return '. '.join(bits) if bits else None

# ---- Prospecting Maint Accounts ----
for r in rows(wb['Prospecting Maint Accounts']):
    co = clean(r.get('Company'))
    if not co: continue
    st = (clean(r.get('Status')) or '').lower()
    add(sheet='maint-prospect', company=co,
        contact=clean(r.get('Contact')), email=clean(r.get('Email')),
        phone=clean(r.get('Phone')), location=clean(r.get('location')),
        type='maintenance', status=MAINT_STATUS.get(st, 'lead'),
        what=clean(r.get('Services')),
        requirement=joined(clean(r.get('Vehicles')), clean(r.get('Requirements'))),
        action=clean(r.get('Update')), next_action=clean(r.get('Next Action')),
        date=when(r.get('Date of update')),
        notes=joined(f"Status on the sheet: {clean(r.get('Status'))}" if r.get('Status') else None,
                     clean(r.get('Update'))),
        source='Prospecting')

# ---- the three trailer sheets ----
for sheet, status in [('Trailer Sales prospects', 'contacted'),
                      ('Trailer Sales Customer', 'customer'),
                      ('Trailer Lost Sales', 'lost')]:
    for r in rows(wb[sheet]):
        co = clean(r.get('Company'))
        if not co: continue
        add(sheet=sheet.lower().replace(' ', '-'), company=co,
            contact=clean(r.get('Contact')), email=clean(r.get('Email')),
            phone=clean(r.get('Phone')), location=None,
            type='trailer_sales', status=status,
            what=clean(r.get('Description')),
            requirement=clean(r.get('Requirement')),
            new_or_used=clean(r.get('col6')),
            estimated_value=money(r.get('Estimated Sales Value')),
            action=clean(r.get('Action')), next_action=None,
            date=when(r.get('Date of Enquiry')),
            notes=joined(clean(r.get('Update')),
                         f"Source: {clean(r.get('Source'))}" if r.get('Source') else None),
            source=clean(r.get('Source')))

# ---- Contract pipeline ----
CONTRACT_TYPE = {'rental/hire': 'rental'}

# FPM and Trukplan are the same maintenance product under two older
# names, and FleetSmart+ replaced both. Left as they were typed, each one
# raises its own filter chip on the maintenance tracker for a product
# nobody sells any more, so somebody filtering by kind of work is offered
# three names for one thing.
#
# The sheet's own word is kept in the notes, so this renames the product
# without losing what the row said. To put them back:
#
#   UPDATE crm_leads SET what = split_part(notes, ': ', 2)
#    WHERE notes LIKE 'Contract pipeline: %';
PRODUCT = {'fpm': 'FleetSmart+', 'trukplan': 'FleetSmart+'}

for r in rows(wb['Contract pipeline']):
    co = clean(r.get('Customer'))
    if not co: continue
    kind = (clean(r.get('Contract type')) or '').lower()
    add(sheet='contract-pipeline', company=co,
        contact=None, email=None, phone=None, location=clean(r.get('Main Depot')),
        type=CONTRACT_TYPE.get(kind, 'maintenance'),
        status='quoted',
        what=PRODUCT.get(kind, clean(r.get('Contract type'))),
        requirement=(f"{clean(r.get('Volume of assets'))} assets"
                     if r.get('Volume of assets') else None),
        estimated_value=money(r.get('Contract Value')),
        action=None, next_action='Get the contract signed.',
        date=when(r.get('Date sent')),
        notes=joined(f"Contract pipeline: {clean(r.get('Contract type'))}",
                     f"Depot {clean(r.get('Main Depot'))}" if clean(r.get('Main Depot')) not in (None, 'N/A') else None),
        source='Contract pipeline')

# ---- Active Rental Accounts ----
for r in rows(wb['Active Rental Accounts']):
    co = clean(r.get('Company'))
    if not co: continue
    add(sheet='rental-active', company=co,
        contact=clean(r.get('Contact')), email=clean(r.get('Email')),
        phone=clean(r.get('Phone')), location=None,
        type='rental', status='customer',
        what=clean(r.get('Asset')),
        requirement=joined(f"{clean(r.get('Quantity'))} on hire" if r.get('Quantity') else None,
                           clean(r.get('Agreement')),
                           f"Service {clean(r.get('Service Cycle'))}" if r.get('Service Cycle') else None),
        estimated_value=None,
        action=None, next_action=None,
        date=when(r.get('Hire start date')),
        notes=joined(f"Weekly rate {clean(r.get('Weekly rate'))}" if r.get('Weekly rate') else None,
                     f"R&M {clean(r.get('R&M'))}, tyres {clean(r.get('Tyres'))}"),
        source='Rental')

print(json.dumps(leads, indent=0)[:400])
print('TOTAL', len(leads))
import collections
print(collections.Counter((l['type'], l['status']) for l in leads))
json.dump(leads, open('/tmp/leads.json','w'), indent=1)
